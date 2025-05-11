"""facemetric.ipynb """

import cv2
import numpy as np
import pandas as pd
from google.colab.patches import cv2_imshow
import dlib

from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as OpenpyxlImage
import tempfile
import os

from google.colab import drive
drive.mount('/content/drive')

cd /content/drive/MyDrive/sample\ data

pwd

!wget http://dlib.net/files/shape_predictor_68_face_landmarks.dat.bz2
!bzip2 -d shape_predictor_68_face_landmarks.dat.bz2

def find_intersection(line1_point1, line1_point2, line2_point1, line2_point2):
    """
    Finds the intersection point of two lines given by two points on each line.

    Parameters:
    - line1_point1, line1_point2: Points on the first line (numpy arrays or tuples).
    - line2_point1, line2_point2: Points on the second line (numpy arrays or tuples).

    Returns:
    - A tuple (x, y) representing the intersection point.
    """
    x1, y1 = line1_point1
    x2, y2 = line1_point2
    x3, y3 = line2_point1
    x4, y4 = line2_point2

    # Calculating the determinant of the system
    den = (x1 - x2) * (y3 - y4) - (y1 - y2) * (x3 - x4)

    # If the lines are parallel (denominator is 0), there is no intersection
    if den == 0:
        return None

    # Calculating the intersection point
    intersect_x = ((x1*y2 - y1*x2)*(x3 - x4) - (x1 - x2)*(x3*y4 - y3*x4)) / den
    intersect_y = ((x1*y2 - y1*x2)*(y3 - y4) - (y1 - y2)*(x3*y4 - y3*x4)) / den

    return (intersect_x, intersect_y)

def calculate_distance(p1, p2):
    """Calculate the Euclidean distance between two points."""
    return np.sqrt((p1[0] - p2[0]) ** 2 + (p1[1] - p2[1]) ** 2)

# Load the pre-trained model for facial landmark detection
predictor = dlib.shape_predictor('shape_predictor_68_face_landmarks.dat')
detector = dlib.get_frontal_face_detector()

# Load the image
image = cv2.imread('yuz_a.jpg')
gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

# Detect faces in the image
faces = detector(gray, 1)

# Initialize the lists to hold point descriptions and their corresponding distances
points = []
distances = []

for face in faces:
    # Determine the facial landmarks for the face region
    shape = predictor(gray, face)

    # Get the coordinates for landmarks 28 and 30
    point_28 = (shape.part(27).x, shape.part(27).y)
    point_31 = (shape.part(30).x, shape.part(30).y)

    cv2.circle(image, (point_28), 3, (255, 0, 0), -1)  # Draws a red dot
    cv2.circle(image, (point_31), 3, (255, 0, 0), -1)  # Draws a red dot

    # Calculate the direction vector of the line
    direction = np.array([point_32[0] - point_28[0], point_31[1] - point_28[1]])

    # Extend the line to the top and bottom of the face
    top_point = np.array(point_28) - direction * 1.5   # Arbitrary large number
    bottom_point = np.array(point_31) + direction * 2.5 # Arbitrary large number

    # Draw the extended line
    cv2.line(image, tuple(top_point.astype(int)), tuple(bottom_point.astype(int)), (255, 0, 0), 2)


    # Calculate and draw symmetry for each facial landmark
    for i in ([0,1,2,3,4, 5, 6, 7]):
        original_point = (shape.part(i).x, shape.part(i).y)

        cv2.circle(image, original_point, 3, (0, 255, 0), -1)
        facial_point = np.array([shape.part(i).x, shape.part(i).y])

        perp_direction = np.array([-direction[1], direction[0]])  # Rotate the direction vector by 90 degrees
        end_point = facial_point - perp_direction * 3
        #cv2.circle(image, end_point, 3, (0,0,255), -1)
        #cv2.line(image, tuple(facial_point.astype(int)), tuple(end_point.astype(int)), (0, 255, 0), 2)

        """Finds the intersection of two lines given in form of ([x1, y1], [x2, y2])."""
        intersect_point  = find_intersection(top_point, bottom_point, facial_point, end_point)
        intersect_point = np.array(intersect_point).astype(int)

        cv2.circle(image, intersect_point, 3, (0,0,255), -1)
        cv2.line(image, tuple(facial_point.astype(int)), tuple(intersect_point), (0, 255, 0), 2)

        distance = calculate_distance(facial_point, intersect_point)
        print(f'Distance between facial point \'{i}\' and its intersection point: {distance}')

        points.append(f'Point {i}')
        distances.append(distance)

    for i in ([9, 10,11,12,13,14,15,16]):
        original_point = (shape.part(i).x, shape.part(i).y)

        cv2.circle(image, original_point, 3, (0, 255, 0), -1)
        facial_point = np.array([shape.part(i).x, shape.part(i).y])

        perp_direction = np.array([direction[1], -direction[0]])  # Rotate the direction vector by 270 degrees
        end_point = facial_point - perp_direction * 3
        #cv2.circle(image, end_point, 3, (0,0,255), -1)
        #cv2.line(image, tuple(end_point.astype(int)), tuple(facial_point.astype(int)), (0, 0, 0), 2)

        intersect_point  = find_intersection(top_point, bottom_point, end_point, facial_point)
        intersect_point = np.array(intersect_point).astype(int)

        cv2.circle(image, intersect_point, 3, (0,0,255), -1)
        cv2.line(image, tuple(intersect_point), tuple(facial_point.astype(int)), (0, 255, 0), 2)

        distance = calculate_distance(facial_point, intersect_point)
        print(f'Distance between facial point \'{i}\' and its intersection point: {distance}')

        points.append(f'Point {i}')
        distances.append(distance)


print(f'points:{points}')
print(f'distance:{distances}')

# Create a DataFrame
df = pd.DataFrame({
    'Point': points,
    'Distance': distances
})

# Save DataFrame to Excel
excel_path = 'distances_and_image.xlsx'
df.to_excel(excel_path, index=False)

# Save your image to a temporary file
temp_image_path = tempfile.mktemp(suffix='.png')
cv2.imwrite(temp_image_path, image)  # Assuming 'image' is your cv2 image

# Load the workbook and get the active sheet
wb = load_workbook(excel_path)
ws = wb.active

# Insert the image into the Excel file
img = OpenpyxlImage(temp_image_path)
# Adjust as needed; 'A' + str(len(distances) + 2) positions the image below the data
img.anchor = 'A' + str(len(distances) + 2)
ws.add_image(img)

# Save the workbook
wb.save(excel_path)

# Optionally, clean up the temporary image file
os.remove(temp_image_path)

# Display the image
cv2_imshow(image)
cv2.waitKey(0)
cv2.destroyAllWindows()

class FacialLandmarkProcessor:
    def __init__(self, predictor_path):
        self.predictor_path = predictor_path
        self.predictor = dlib.shape_predictor(predictor_path)
        self.detector = dlib.get_frontal_face_detector()

    def find_intersection(self, line1_point1, line1_point2, line2_point1, line2_point2):
        x1, y1 = line1_point1
        x2, y2 = line1_point2
        x3, y3 = line2_point1
        x4, y4 = line2_point2

        # Calculating the determinant of the system
        den = (x1 - x2) * (y3 - y4) - (y1 - y2) * (x3 - x4)

        # If the lines are parallel (denominator is 0), there is no intersection
        if den == 0:
            return None

        # Calculating the intersection point
        intersect_x = ((x1*y2 - y1*x2)*(x3 - x4) - (x1 - x2)*(x3*y4 - y3*x4)) / den
        intersect_y = ((x1*y2 - y1*x2)*(y3 - y4) - (y1 - y2)*(x3*y4 - y3*x4)) / den

        return (intersect_x, intersect_y)


    def calculate_distance(self, p1, p2):
        """Calculate the Euclidean distance between two points."""
        return np.sqrt((p1[0] - p2[0]) ** 2 + (p1[1] - p2[1]) ** 2)

    def process_image(self, image_path):
        image = cv2.imread(image_path)
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        faces = self.detector(gray, 1)

        points = []
        distances = []


        for face in faces:
            shape = self.predictor(gray, face)
            # Get the coordinates for landmarks 28 and 30
            point_28 = (shape.part(27).x, shape.part(27).y)
            point_31 = (shape.part(30).x, shape.part(30).y)

            cv2.circle(image, (point_28), 3, (255, 0, 0), -1)  # Draws a red dot
            cv2.circle(image, (point_31), 3, (255, 0, 0), -1)  # Draws a red dot

            # Calculate the direction vector of the line
            direction = np.array([point_31[0] - point_28[0], point_31[1] - point_28[1]])

            # Extend the line to the top and bottom of the face
            top_point = np.array(point_28) - direction * 1.5   # Arbitrary large number
            bottom_point = np.array(point_31) + direction * 2.5 # Arbitrary large number

            # Draw the extended line
            cv2.line(image, tuple(top_point.astype(int)), tuple(bottom_point.astype(int)), (255, 0, 0), 2)


            # Calculate and draw symmetry for each facial landmark
            for i in ([0,1,2,3,4, 5, 6, 7]):
                original_point = (shape.part(i).x, shape.part(i).y)

                cv2.circle(image, original_point, 3, (0, 255, 0), -1)
                facial_point = np.array([shape.part(i).x, shape.part(i).y])

                perp_direction = np.array([-direction[1], direction[0]])  # Rotate the direction vector by 90 degrees
                end_point = facial_point - perp_direction * 3
                #cv2.circle(image, end_point, 3, (0,0,255), -1)
                #cv2.line(image, tuple(facial_point.astype(int)), tuple(end_point.astype(int)), (0, 255, 0), 2)

                """Finds the intersection of two lines given in form of ([x1, y1], [x2, y2])."""
                intersect_point  = find_intersection(top_point, bottom_point, facial_point, end_point)
                intersect_point = np.array(intersect_point).astype(int)

                cv2.circle(image, intersect_point, 3, (0,0,255), -1)
                cv2.line(image, tuple(facial_point.astype(int)), tuple(intersect_point), (0, 255, 0), 2)

                distance = calculate_distance(facial_point, intersect_point)
                print(f'Distance between facial point \'{i}\' and its intersection point: {distance}')

                points.append(f'Point {i}')
                distances.append(distance)

            for i in ([9, 10,11,12,13,14,15,16]):
                original_point = (shape.part(i).x, shape.part(i).y)

                cv2.circle(image, original_point, 3, (0, 255, 0), -1)
                facial_point = np.array([shape.part(i).x, shape.part(i).y])

                perp_direction = np.array([direction[1], -direction[0]])  # Rotate the direction vector by 270 degrees
                end_point = facial_point - perp_direction * 3
                #cv2.circle(image, end_point, 3, (0,0,255), -1)
                #cv2.line(image, tuple(end_point.astype(int)), tuple(facial_point.astype(int)), (0, 0, 0), 2)

                intersect_point  = find_intersection(top_point, bottom_point, end_point, facial_point)
                intersect_point = np.array(intersect_point).astype(int)

                cv2.circle(image, intersect_point, 3, (0,0,255), -1)
                cv2.line(image, tuple(intersect_point), tuple(facial_point.astype(int)), (0, 255, 0), 2)

                distance = calculate_distance(facial_point, intersect_point)
                print(f'Distance between facial point \'{i}\' and its intersection point: {distance}')

                points.append(f'Point {i}')
                distances.append(distance)
        print(len(distances))

        # Return the modified image, points list, and distances list
        return image, points, distances


    def calculate_symetry(self, distances):
        distances = np.array(distances)

        # Calculate symmetry using vectorized operations
        symmetries = (1 - distances[:8] / distances[15:7:-1]) ** 2
        symmetry_metric = np.mean(symmetries)

        return symmetry_metric


    def save_data_to_excel(self, image_paths, excel_path):
        wb = load_workbook(excel_path) if os.path.exists(excel_path) else Workbook()
        for idx, image_path in enumerate(image_paths):
            image, points, distances = self.process_image(image_path)

            symetry_metric = self.calculate_symetry(distances)

            ws = wb.create_sheet(f"Face {idx + 1}")
            df = pd.DataFrame({'Point': points, 'Distance': distances, 'Symetry_Value': symetry_metric1})
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            temp_image_path = tempfile.mktemp(suffix='.png')
            cv2.imwrite(temp_image_path, image)
            img = OpenpyxlImage(temp_image_path)
            img.anchor = 'A' + str(len(distances) + 2)
            ws.add_image(img)

        wb.save(excel_path)


# Example usage
predictor_path = 'shape_predictor_68_face_landmarks.dat'
processor = FacialLandmarkProcessor(predictor_path)
image_paths = ['yuz_a.jpg', 'yuz_b.jpg']
excel_path = 'distances_and_image.xlsx'
processor.save_data_to_excel(image_paths, excel_path)

